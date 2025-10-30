# # ai_model_server.py - Simple API for your AI model
# from flask import Flask, request, jsonify, send_file
# import pandas as pd
# import numpy as np
# import joblib
# import os
# from werkzeug.utils import secure_filename
# import uuid
# from datetime import datetime

# app = Flask(__name__)

# # Configuration
# UPLOAD_FOLDER = 'uploads'
# RESULTS_FOLDER = 'results'
# MODEL_PATH = 'risk_model.pkl'

# # Create directories
# os.makedirs(UPLOAD_FOLDER, exist_ok=True)
# os.makedirs(RESULTS_FOLDER, exist_ok=True)

# # Load model once at startup
# print("Loading AI model...")
# try:
#     model_data = joblib.load(MODEL_PATH)
#     MODEL = model_data['model']
#     SCALER = model_data['scaler']
#     FEATURE_COLUMNS = model_data['feature_columns']
#     print("✅ Model loaded successfully!")
# except Exception as e:
#     print(f"❌ Error loading model: {e}")
#     MODEL = None

# @app.route('/')
# def home():
#     return jsonify({
#         'message': 'Risk Classification AI API',
#         'status': 'Model loaded' if MODEL else 'Model not found',
#         'endpoints': {
#             '/predict': 'POST - Upload Excel file for risk prediction'
#         }
#     })

# @app.route('/predict', methods=['POST'])
# def predict():
#     try:
#         # Check if model is loaded
#         if MODEL is None:
#             return jsonify({
#                 'success': False,
#                 'error': 'AI model not loaded'
#             }), 500

#         # Check if file is uploaded
#         if 'file' not in request.files:
#             return jsonify({
#                 'success': False,
#                 'error': 'No file uploaded'
#             }), 400

#         file = request.files['file']
#         if file.filename == '':
#             return jsonify({
#                 'success': False,
#                 'error': 'No file selected'
#             }), 400

#         # Save uploaded file
#         filename = secure_filename(file.filename)
#         unique_id = str(uuid.uuid4())[:8]
#         input_filename = f"{unique_id}_{filename}"
#         input_path = os.path.join(UPLOAD_FOLDER, input_filename)
#         file.save(input_path)

#         # Read Excel file
#         df = pd.read_excel(input_path)
#         print(f"Input file: {len(df)} rows, {len(df.columns)} columns")

#         # Prepare features
#         for col in FEATURE_COLUMNS:
#             if col not in df.columns:
#                 df[col] = 0  # Add missing columns

#         X = df[FEATURE_COLUMNS].fillna(0)

#         # Make predictions
#         X_scaled = SCALER.transform(X)
#         predictions = MODEL.predict(X_scaled)
#         probabilities = MODEL.predict_proba(X_scaled)

#         # Add predictions to original dataframe
#         df['risk_bucket'] = predictions
#         df['prediction_confidence'] = np.max(probabilities, axis=1)

#         # Generate output filename
#         output_filename = f"result_{unique_id}_{filename}"
#         output_path = os.path.join(RESULTS_FOLDER, output_filename)

#         # Save result Excel
#         df.to_excel(output_path, index=False)

#         # Calculate stats
#         prediction_stats = {
#             'Low': int(np.sum(predictions == 'Low')),
#             'Medium': int(np.sum(predictions == 'Medium')),
#             'High': int(np.sum(predictions == 'High'))
#         }

#         # Clean up input file
#         os.remove(input_path)

#         return jsonify({
#             'success': True,
#             'message': 'Risk prediction completed',
#             'total_records': len(df),
#             'predictions': prediction_stats,
#             'result_filename': output_filename,
#             'download_url': f'/download/{output_filename}'
#         })

#     except Exception as e:
#         print(f"Error in prediction: {e}")
#         # Clean up files
#         if os.path.exists(input_path):
#             os.remove(input_path)
        
#         return jsonify({
#             'success': False,
#             'error': str(e)
#         }), 500

# @app.route('/download/<filename>')
# def download_file(filename):
#     try:
#         file_path = os.path.join(RESULTS_FOLDER, filename)
#         if os.path.exists(file_path):
#             return send_file(file_path, as_attachment=True)
#         else:
#             return jsonify({
#                 'success': False,
#                 'error': 'File not found'
#             }), 404
#     except Exception as e:
#         return jsonify({
#             'success': False,
#             'error': str(e)
#         }), 500

# if __name__ == '__main__':
#     print("🚀 Starting AI Model Server...")
#     print("📁 Make sure 'risk_model.pkl' is in the same directory")
#     app.run(debug=True, host='0.0.0.0', port=5000)

# ai_model_server_no_pandas_numpy.py - Simple API for your AI model without pandas/numpy
# from flask import Flask, request, jsonify, send_file
# import joblib
# import os
# from werkzeug.utils import secure_filename
# import uuid
# from openpyxl import load_workbook, Workbook

# app = Flask(__name__)

# # Configuration
# UPLOAD_FOLDER = 'uploads'
# RESULTS_FOLDER = 'results'
# MODEL_PATH = 'risk_model.pkl'

# # Create directories
# os.makedirs(UPLOAD_FOLDER, exist_ok=True)
# os.makedirs(RESULTS_FOLDER, exist_ok=True)

# # Load model once at startup
# print("Loading AI model...")
# try:
#     model_data = joblib.load(MODEL_PATH)
#     MODEL = model_data['model']
#     SCALER = model_data['scaler']
#     FEATURE_COLUMNS = model_data['feature_columns']
#     print("✅ Model loaded successfully!")
# except Exception as e:
#     print(f"❌ Error loading model: {e}")
#     MODEL = None

# @app.route('/')
# def home():
#     return jsonify({
#         'message': 'Risk Classification AI API',
#         'status': 'Model loaded' if MODEL else 'Model not found',
#         'endpoints': {
#             '/predict': 'POST - Upload Excel file for risk prediction'
#         }
#     })

# @app.route('/predict', methods=['POST'])
# def predict():
#     try:
#         # if MODEL is None:
#         #     return jsonify({'success': False, 'error': 'AI model not loaded'}), 500

#         # if 'file' not in request.files:
#         #     return jsonify({'success': False, 'error': 'No file uploaded'}), 400

#         file = request.files['file']
#         if file.filename == '':
#             return jsonify({'success': False, 'error': 'No file selected'}), 400

#         # Save uploaded file
#         filename = secure_filename(file.filename)
#         unique_id = str(uuid.uuid4())[:8]
#         input_filename = f"{unique_id}_{filename}"
#         input_path = os.path.join(UPLOAD_FOLDER, input_filename)
#         file.save(input_path)

#         # Read Excel using openpyxl
#         wb = load_workbook(input_path)
#         ws = wb.active

#         headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
#         data_rows = list(ws.iter_rows(min_row=2, values_only=True))

#         # Prepare feature matrix
#         X = []
#         for row in data_rows:
#             row_dict = {headers[i]: row[i] if i < len(row) else 0 for i in range(len(headers))}
#             # Add missing feature columns
#             row_features = [row_dict.get(col, 0) for col in FEATURE_COLUMNS]
#             X.append(row_features)

#         # Scale features and make predictions
#         X_scaled = SCALER.transform(X)
#         predictions = MODEL.predict(X_scaled)
#         probabilities = MODEL.predict_proba(X_scaled)

#         # Add predictions back to Excel
#         ws.cell(row=1, column=len(headers)+1, value='risk_bucket')
#         ws.cell(row=1, column=len(headers)+2, value='prediction_confidence')

#         for idx, (pred, prob) in enumerate(zip(predictions, probabilities)):
#             ws.cell(row=idx+2, column=len(headers)+1, value=pred)
#             ws.cell(row=idx+2, column=len(headers)+2, value=float(max(prob)))

#         # Save result Excel
#         output_filename = f"result_{unique_id}_{filename}"
#         output_path = os.path.join(RESULTS_FOLDER, output_filename)
#         wb.save(output_path)

#         # Stats calculation
#         prediction_stats = {'Low': 0, 'Medium': 0, 'High': 0}
#         for p in predictions:
#             if p in prediction_stats:
#                 prediction_stats[p] += 1

#         os.remove(input_path)  # cleanup

#         return jsonify({
#             'success': True,
#             'message': 'Risk prediction completed',
#             'total_records': len(data_rows),
#             'predictions': prediction_stats,
#             'result_filename': output_filename,
#             'download_url': f'/download/{output_filename}'
#         })

#     except Exception as e:
#         if os.path.exists(input_path):
#             os.remove(input_path)
#         return jsonify({'success': False, 'error': str(e)}), 500

# @app.route('/download/<filename>')
# def download_file(filename):
#     try:
#         file_path = os.path.join(RESULTS_FOLDER, filename)
#         if os.path.exists(file_path):
#             return send_file(file_path, as_attachment=True)
#         else:
#             return jsonify({'success': False, 'error': 'File not found'}), 404
#     except Exception as e:
#         return jsonify({'success': False, 'error': str(e)}), 500

# if __name__ == '__main__':
#     print("🚀 Starting AI Model Server...")
#     print("📁 Make sure 'risk_model.pkl' is in the same directory")
#     app.run(debug=True, host='0.0.0.0', port=5000)
# ai_model_server.py - AI model API supporting Excel and CSV
from flask import Flask, request, jsonify, send_file
import joblib
import os
from werkzeug.utils import secure_filename
import uuid
from openpyxl import load_workbook, Workbook
import csv

app = Flask(__name__)

# Configuration
UPLOAD_FOLDER = 'uploads'
RESULTS_FOLDER = 'results'
MODEL_PATH = 'risk_model.pkl'

# Create directories
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULTS_FOLDER, exist_ok=True)

# Load model once at startup
print("Loading AI model...")
try:
    model_data = joblib.load(MODEL_PATH)
    MODEL = model_data['model']
    SCALER = model_data['scaler']
    FEATURE_COLUMNS = model_data['feature_columns']
    print("✅ Model loaded successfully!")
except Exception as e:
    print(f"❌ Error loading model: {e}")
    MODEL = None

@app.route('/')
def home():
    return jsonify({
        'message': 'Risk Classification AI API',
        'status': 'Model loaded' if MODEL else 'Model not found',
        'endpoints': {
            '/predict': 'POST - Upload Excel (.xlsx) or CSV (.csv) file for risk prediction'
        }
    })

@app.route('/predict', methods=['POST'])
def predict():
    try:
        if MODEL is None:
            return jsonify({'success': False, 'error': 'AI model not loaded'}), 500

        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400

        # Save uploaded file
        filename = secure_filename(file.filename)
        unique_id = str(uuid.uuid4())[:8]
        input_filename = f"{unique_id}_{filename}"
        input_path = os.path.join(UPLOAD_FOLDER, input_filename)
        file.save(input_path)

        # Read file based on type
        if filename.endswith('.xlsx'):
            wb = load_workbook(input_path)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        elif filename.endswith('.csv'):
            with open(input_path, 'r', newline='', encoding='utf-8') as f:
                reader = csv.reader(f)
                headers = next(reader)
                data_rows = list(reader)
        else:
            return jsonify({'success': False, 'error': 'Unsupported file type'}), 400

        # Prepare feature matrix
        X = []
        for row in data_rows:
            row_dict = {headers[i]: row[i] if i < len(row) else 0 for i in range(len(headers))}
            row_features = [row_dict.get(col, 0) for col in FEATURE_COLUMNS]
            X.append(row_features)

        # Scale and predict
        X_scaled = SCALER.transform(X)
        predictions = MODEL.predict(X_scaled)
        probabilities = MODEL.predict_proba(X_scaled)

        # Add predictions back to output
        if filename.endswith('.xlsx'):
            ws.cell(row=1, column=len(headers)+1, value='risk_bucket')
            ws.cell(row=1, column=len(headers)+2, value='prediction_confidence')
            for idx, (pred, prob) in enumerate(zip(predictions, probabilities)):
                ws.cell(row=idx+2, column=len(headers)+1, value=pred)
                ws.cell(row=idx+2, column=len(headers)+2, value=float(max(prob)))
            output_path = os.path.join(RESULTS_FOLDER, f"result_{unique_id}_{filename}")
            wb.save(output_path)
        else:  # CSV
            output_path = os.path.join(RESULTS_FOLDER, f"result_{unique_id}_{filename}")
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(headers + ['risk_bucket', 'prediction_confidence'])
                for row, pred, prob in zip(data_rows, predictions, probabilities):
                    writer.writerow(list(row) + [pred, float(max(prob))])

        # Stats calculation
        prediction_stats = {'Low': 0, 'Medium': 0, 'High': 0}
        for p in predictions:
            if p in prediction_stats:
                prediction_stats[p] += 1

        os.remove(input_path)  # cleanup

        return jsonify({
            'success': True,
            'message': 'Risk prediction completed',
            'total_records': len(data_rows),
            'predictions': prediction_stats,
            'result_filename': os.path.basename(output_path),
            'download_url': f'/download/{os.path.basename(output_path)}'
        })

    except Exception as e:
        if os.path.exists(input_path):
            os.remove(input_path)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/download/<filename>')
def download_file(filename):
    try:
        file_path = os.path.join(RESULTS_FOLDER, filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True)
        else:
            return jsonify({'success': False, 'error': 'File not found'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    print("🚀 Starting AI Model Server...")
    print("📁 Make sure 'risk_model.pkl' is in the same directory")
    app.run(debug=True, host='0.0.0.0', port=5000)




